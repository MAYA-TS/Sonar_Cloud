import xlsxwriter
import os
import pandas as pd
import cx_Oracle
import openpyxl
from pandas import ExcelWriter
from openpyxl import Workbook
from email.message import EmailMessage
from email.utils import make_msgid
import mimetypes
from tableau_api_lib import TableauServerConnection
from tableau_api_lib.utils import querying
from tableau_api_lib.utils.common import flatten_dict_column
import pandas as pd
import xlrd
from pandas import ExcelWriter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
import smtplib
import datetime
import time
# import schedule
import time


conn = cx_Oracle.connect("kpmg", "Asd$1234", "HISTDB1")
tableau_server_config = {
                'my_env': {
                        'server': 'https://reports.manappuram.com',
                        'api_version': "3.17",
                        'username': 'tableauadministrator',
                        'password': 'M@fil@123',
                        'site_name': 'Default',
                        'site_url': ''
                }
        }
conn = TableauServerConnection(tableau_server_config, env='my_env',ssl_verify=False)
conn.sign_in()
site_views_df = querying.get_views_dataframe(conn)
site_views_detailed_df = flatten_dict_column(site_views_df, keys=['name', 'id'], col_name='workbook')
site_views_detailed_df.tail(60)
relevant_views_df = site_views_detailed_df[site_views_detailed_df['workbook_name'] == 'Gold photo verification Report']
print(relevant_views_df)
# relevant_views_df.refresh(workbook_id='949d788d-c89f-432d-8af2-c03990e149c5')

# fzm_data='4bb1459c-b3e6-4a5b-8b07-0eaf77f98a36'
fzm_data='8a45b7db-e53e-4637-a173-777a7aecc370'
view_img = conn.query_view_image(view_id=fzm_data)
print(view_img)
with open(r"C:\\Users\\398504\\CRF\\crf23\\Gold_photo_verification_122016\\Gold_verification_report.png","wb") as f:
   
    f.write(view_img.content)
print("Third section completed................")
conn = cx_Oracle.connect("kpmg", "Asd$1234", "HISTDB1")
print("Oracle database connected")
print("Tableau server connected")

        


df1=pd.read_sql("""select distinct q.cust_id,
                e.name customer_name,
                q.branch_id,
                q.pledge_no,
                q.pledge_val,
                z.area_name,
                z.reg_name,
                z.zonal_name

  from mana0809.pledge_master@uatr_backup2 q
  left outer join mana0809.customer@uatr_backup2 e
    on (e.cust_id = q.cust_id)
  LEFT OUTER JOIN tbl_fake_gold_verification s
    ON (q.pledge_no = s.pledge_no)
  left outer join mana0809.branch_dtl_new@uatr_backup2 z
    on (q.branch_id = z.branch_id)
 where s.status = 1
   AND s.CREATED_DT = trunc(sysdate) - 1""",con=conn)
#---------------------------------------------------------------------


print("Query section completed...........")



if not df1.empty:
    writer=pd.ExcelWriter("C:\\Users\\398504\\CRF\\crf23\\Gold_photo_verification_122016\\Gold_verification_report.xlsx",engine="openpyxl")
    df1.to_excel(writer, sheet_name="Gold_verification_report", index=False)


    print("saved as excel")
    writer.save()
    print("Excel downloading section completed............")


    workbook = openpyxl.load_workbook('C:\\Users\\398504\\CRF\\crf23\\Gold_photo_verification_122016\\Gold_verification_report.xlsx')
    print("wb open")
    sheet_names = workbook.sheetnames
    heading_color =  'E2BCB7'    #'73BCC5'#'8080ff'  # Red color
    # body_color = 'F6E5F5'  # Green color
    border_style = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin'))


    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        header_font = Font(color="000000", bold=True)
        header_fill = PatternFill(start_color='E2BCB7', end_color='E2BCB7', fill_type='solid')
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        # body_fill = PatternFill(start_color=body_color, end_color=body_color, fill_type='solid')
        # for row in sheet.iter_rows(min_row=2):
        #     for cell in row:
        #         cell.fill = body_fill
                
        for column in sheet.columns:
            non_empty_values = [cell.value for cell in column if cell.value]
            if non_empty_values:
                max_length = max(len(str(value)) for value in non_empty_values)
                column_letter = get_column_letter(column[0].column)
                adjusted_width = (max_length + 2) * 1.2  # Adjust the width as desired
                sheet.column_dimensions[column_letter].width = adjusted_width
        for row in sheet.rows:
            max_height = max(str(cell.value).count('\n') + 1 for cell in row if cell.value)
            row_number = row[0].row
            adjusted_height = max_height * 17 # Adjust the height as desired
            sheet.row_dimensions[row_number].height = adjusted_height
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border_style
    workbook.save('C:\\Users\\398504\\CRF\\crf23\\Gold_photo_verification_122016\\Gold_verification_report.xlsx')
    print("done")
   
    workbook = openpyxl.load_workbook(r"C:\\Users\\398504\\CRF\\crf23\\Gold_photo_verification_122016\\Gold_verification_report.xlsx")
    worksheet = workbook["Gold_verification_report"]

    # if worksheet.max_row == 0:
    s = smtplib.SMTP(host='smtp.office365.com', port=587)
    s.starttls()
    # print("aaa")
  
    s.login('iotautomation@manappuram.com','ybjmxbfdyzkdnjtw')
    
    print("Login the mail address")
    msg = EmailMessage()
    print("Ready for mailing")

    msg['Subject'] = 'Gold ornament photos not available for yesterday pledges'
    # msg['From'] = '<internalaudit1@manappuram.com>'
    # msg['To'] = 'Audit Research Wing<researchwing@manappuram.com>','LAXMAN TAGGINAVAR <headresearchwing@manappuram.com>'
    # msg['Cc'] = 'RIJU P<gmaudit@manappuram.com>','ANN MARY M B<dataservice26@manappuram.com>','ANJANA V P<iotsupport15@manappuram.com>'

    msg['From']='<iotautomation@manappuram.com>'
    # msg['To']='<glalertsaudit@manappuram.com>'

    with open(r"C:\\Users\\398504\\CRF\\crf23\\Gold_photo_verification_122016\\Gold_verification_report.xlsx", 'rb') as ra:
        attachment = ra.read()
    msg.add_related(attachment, maintype='application', subtype='xlsx', filename='Gold verification report.xlsx')
    image_cid = make_msgid(domain='mandala.com')
    msg.add_alternative("""\
    <html>
        <body>
            <p>Dear Sir,<br><p/>      
            <p>Please find the attachment. </p>
                        <p>
                        <p>
            <P>  <img src="cid:{image_cid}">
                        </>      
            <p>
            <p> 
            <p>
                Thanks & Regards,<br>
                Team IoT <br>
                R & D New age technology <br>
                ( This is an autogenerated mail )
            </p>
        </body>
    </html>
    """.format(image_cid=image_cid[1:-1]),subtype='html')    



    with open(r"C:\\Users\\398504\\CRF\\crf23\\Gold_photo_verification_122016\\Gold_verification_report.png", 'rb') as img:
        maintype, subtype = mimetypes.guess_type(img.name)[0].split('/')
        msg.get_payload()[1].add_related(img.read(),
                                            maintype=maintype,
                                            subtype=subtype,
                                            cid=image_cid)


    s.send_message(msg)
    s.quit()
    print("Mail send")
    print("final section completed sucessfully.............")
    # os.remove(r"C:\\Users\\398504\\CRF\\crf22\\Fake_NEFT_Verification-121681\\fake_NEFT_report.png")
    print("Image removed")
else:
    print("                                                                     ")
    print("                                                                     ")
    print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!NO DATA IN DB TABLE!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
        
